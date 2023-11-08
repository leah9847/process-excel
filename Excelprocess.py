import openpyxl
import os
import sys
import time

def process_files(folder, out_file):
	count=0
	start = time.time()
	if out_file.endswith(".xls") or out_file.endswith(".xlsx"):
		wb_out = openpyxl.load_workbook(outfile)
	else:
		print("Invalid output file, not Excel spreadsheet")
		return
	ws_out = wb_out.active

	files = (file for file in os.listdir(folder) if not os.path.isdir(file))

	#loop through all of the input files
	for file in files:
		if file.endswith(".xls") or file.endswith(".xlsx"):
			wb = openpyxl.load_workbook(file)
			ifile = os.path.join(folder, file)
			ws = wb.active #get the first worksheet
			
			#get the specific cells that I want
			
			sa_length = ws['G151'].value
			pln=[ws['G155'].value, ws['G159'].value, ws['G163'].value, ws['G167'].value, ws['G171'].value, ws['G175'].value, ws['G179'].value, ws['G183'].value, ws['G187'].value, ws['G191'].value,
				ws['G195'].value, ws['G199'].value, ws['G203'].value, ws['G207'].value, ws['G211'].value, ws['G215'].value, ws['G219'].value, ws['G223'].value, ws['G227'].value, ws['G231'].value,
				ws['G235'].value, ws['G239'].value, ws['G243'].value, ws['G247'].value, ws['G251'].value, ws['G255'].value, ws['G259'].value, ws['G263'].value, ws['G267'].value
				]
			pln_max = ws['G275'].value
			rect=[ws['G283'].value, ws['G287'].value, ws['G291'].value, ws['G295'].value]
			edge1=[ws['G305'].value, ws['G309'].value, ws['G313'].value, ws['G317'].value, ws['G321'].value, ws['G325'].value, ws['G329'].value]
			middle=[ws['G337'].value, ws['G341'].value, ws['G345'].value, ws['G349'].value, ws['G353'].value, ws['G357'].value, ws['G361'].value]
			edge2=[ws['G369'].value, ws['G373'].value, ws['G377'].value, ws['G381'].value, ws['G385'].value, ws['G389'].value, ws['G393'].value]
			new_row = [' ', sa_length, pln, pln_max, rect, edge1, middle, edge2]
			#add a new row to the output excel file
			ws_out.append(new_row)
			#close the input file
			wb.close()
		 
	end = time.time()
	wb_out.save(outfile)
	wb_out.close()
	#total time taken to process the file
	print("Execution time in seconds: ",(end - start))
	
if __name__ == "__main__":
	inputFolder = sys.argv[1]
	outputFile = sys.argv[2]
	print ("input folder ", inputFolder, " output file ", outputFile)

# Check if path exists
	if os.path.exists(inputFolder):
		print ("Input folder OK")
	else:
		sys.exit ("Input folder is required. A folder with embedded spaces must be surrounded by double quotes")
# Check if file exists
	if os.path.isfile(outputFile) 
		print ("output file OK")
	process_files(inputFolder, outputFile) 
